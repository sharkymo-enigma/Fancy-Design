#!/usr/bin/env python3
"""
Convert .xlsx, .docx, and .pdf in knowledge-base/ to .md for the analyst.
Run as hook before the perpetual KB trigger (conversation start).
"""
from pathlib import Path
import sys
from typing import Optional

# --- xlsx (openpyxl) ---
try:
    import openpyxl
except ImportError:
    openpyxl = None


def sanitize_cell(value):
    if value is None:
        return ""
    s = str(value).strip().replace("|", "\\|").replace("\n", " ")
    return s


def sheet_to_markdown(ws, max_rows=2000):
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return ""
    while rows and all(c is None or str(c).strip() == "" for c in (rows[0] or [])):
        rows.pop(0)
    if not rows:
        return ""
    header = list(rows[0]) if rows[0] else []
    data = rows[1:][:max_rows]
    data_lens = [len(r) if r else 0 for r in data]
    col_count = max([len(header)] + data_lens) if data_lens else len(header)
    header = header + [None] * (col_count - len(header))
    header = [sanitize_cell(c) for c in header]
    sep = "| " + " | ".join("---" for _ in header) + " |"
    lines = ["| " + " | ".join(header) + " |", sep]
    for row in data:
        row = list(row) if row else []
        row = row + [None] * (col_count - len(row))
        lines.append("| " + " | ".join(sanitize_cell(c) for c in row) + " |")
    return "\n".join(lines)


def convert_xlsx(xlsx_path: Path, md_path: Path) -> None:
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    parts = [f"# {md_path.stem}\n\nConverted from `{xlsx_path.name}`.\n"]
    for name in wb.sheetnames:
        ws = wb[name]
        table = sheet_to_markdown(ws)
        if table:
            parts.append(f"## Sheet: {name}\n\n{table}\n")
    wb.close()
    md_path.parent.mkdir(parents=True, exist_ok=True)
    md_path.write_text("\n".join(parts), encoding="utf-8")
    print(f"  xlsx -> {md_path.name}")


# --- docx (python-docx) ---
try:
    from docx import Document
    from docx.oxml.table import CT_Tbl
    from docx.oxml.text.paragraph import CT_P
    from docx.table import Table
    from docx.text.paragraph import Paragraph
except ImportError:
    Document = None
    Table = None
    Paragraph = None


def docx_heading_level(style_name: str) -> Optional[int]:
    if not style_name:
        return None
    if style_name == "Title":
        return 1
    if style_name.startswith("Heading"):
        try:
            return int(style_name[-1])
        except ValueError:
            return 1
    return None


def docx_table_to_md(table) -> str:
    rows = []
    for row in table.rows:
        cells = [sanitize_cell(c.text) for c in row.cells]
        rows.append("| " + " | ".join(cells) + " |")
    if not rows:
        return ""
    ncols = len(rows[0].split("|")) - 2
    sep = "| " + " | ".join("---" for _ in range(max(ncols, 1))) + " |"
    return "\n".join([rows[0], sep] + rows[1:])


def iter_block_items(doc):
    parent_elm = doc.element.body
    for child in parent_elm.iterchildren():
        if child.tag.endswith("}p") or child.tag == "p":
            yield ("p", Paragraph(child, doc))
        elif child.tag.endswith("}tbl") or child.tag == "tbl":
            yield ("tbl", Table(child, doc))


def convert_docx(docx_path: Path, md_path: Path) -> None:
    doc = Document(docx_path)
    parts = [f"# {md_path.stem}\n\nConverted from `{docx_path.name}`.\n\n---\n"]
    for kind, block in iter_block_items(doc):
        if kind == "p":
            level = docx_heading_level(block.style.name if block.style else None)
            if level:
                parts.append(f"\n{'#' * level} {block.text}\n")
            elif block.text.strip():
                parts.append(block.text + "\n")
        elif kind == "tbl":
            parts.append("\n" + docx_table_to_md(block) + "\n")
    md_path.parent.mkdir(parents=True, exist_ok=True)
    md_path.write_text("\n".join(parts), encoding="utf-8")
    print(f"  docx -> {md_path.name}")


# --- pdf (pymupdf4llm or pymupdf) ---
try:
    import pymupdf4llm
    _pdf_backend = "pymupdf4llm"
except ImportError:
    try:
        import fitz  # pymupdf
        _pdf_backend = "pymupdf"
    except ImportError:
        _pdf_backend = None


def convert_pdf(pdf_path: Path, md_path: Path) -> None:
    if _pdf_backend == "pymupdf4llm":
        md_text = pymupdf4llm.to_markdown(str(pdf_path))
        body = md_text.strip()
    elif _pdf_backend == "pymupdf":
        doc = fitz.open(pdf_path)
        parts = []
        for page in doc:
            parts.append(page.get_text())
        doc.close()
        body = "\n\n".join(parts).strip()
    else:
        raise RuntimeError("No PDF library: pip install pymupdf4llm or pymupdf")
    content = f"# {md_path.stem}\n\nConverted from `{pdf_path.name}`.\n\n---\n\n{body}\n"
    md_path.parent.mkdir(parents=True, exist_ok=True)
    md_path.write_text(content, encoding="utf-8")
    print(f"  pdf  -> {md_path.name}")


def get_kb_path(root: Path) -> Optional[Path]:
    kb = root / "knowledge-base"
    if kb.exists():
        return kb.resolve()
    path_file = root / ".knowledge-base-path"
    if path_file.exists():
        return Path(path_file.read_text().strip())
    return None


def main():
    root = Path(__file__).resolve().parent.parent
    kb = get_kb_path(root)
    if not kb or not kb.is_dir():
        print("knowledge-base not found.", file=sys.stderr)
        sys.exit(1)

    did_any = False

    if openpyxl:
        for p in sorted(kb.glob("*.xlsx")):
            try:
                convert_xlsx(p, p.with_suffix(".md"))
                did_any = True
            except Exception as e:
                print(f"  ERROR {p.name}: {e}", file=sys.stderr)
    else:
        print("  (skip xlsx: pip install openpyxl)", file=sys.stderr)

    if Document:
        for p in sorted(kb.glob("*.docx")):
            try:
                convert_docx(p, p.with_suffix(".md"))
                did_any = True
            except Exception as e:
                print(f"  ERROR {p.name}: {e}", file=sys.stderr)
    else:
        print("  (skip docx: pip install python-docx)", file=sys.stderr)

    if _pdf_backend:
        for p in sorted(kb.glob("*.pdf")):
            try:
                convert_pdf(p, p.with_suffix(".md"))
                did_any = True
            except Exception as e:
                print(f"  ERROR {p.name}: {e}", file=sys.stderr)
    else:
        print("  (skip pdf: pip install pymupdf4llm or pymupdf)", file=sys.stderr)

    if did_any:
        print("KB artefacts (.xlsx, .docx, .pdf) -> .md done.")
    else:
        print("No .xlsx/.docx/.pdf in knowledge-base, or converters not installed.")


if __name__ == "__main__":
    main()
